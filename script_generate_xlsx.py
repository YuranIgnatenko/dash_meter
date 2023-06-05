import sys
from openpyxl import Workbook, load_workbook
import random, os, datetime
import backend.extract as extr

COUNT_RECORDS = 200
MIN_VALUE = 0
MAX_VALUE = 100
MIN_YEAR = 2017
MAX_YEAR = 2023
NAME_FILE_RANDOM = "test.xlsx"
NAME_FILE_SORTED = "test.xlsx"
NAME_FILE_IMAGES = "assets.xlsx"
PREFIX_RANDOM_IMG = "image_random_"

# перезапись и.или создание
# таблицы со случайными данными
def generate_random_xlsx():
            wb = Workbook()
            table_xlsx = wb.active
            table_xlsx.title = NAME_FILE_RANDOM
            for i in range(COUNT_RECORDS):
                index = str(i+1)
                table_xlsx['A'+index] = PREFIX_RANDOM_IMG+index+".jpg"
                table_xlsx['B'+index] = random.randint(MIN_VALUE, MAX_VALUE)
                rand_year = str(random.randint(MIN_YEAR, MAX_YEAR))
                rand_mount = str(random.randint(1,12))
                rand_day = str(random.randint(1,28))
                if len(rand_mount) == 1: rand_mount = "0"+rand_mount
                if len(rand_day) == 1: rand_day = "0"+rand_day
                table_xlsx['C'+index] = f"{rand_year}-{rand_mount}-{rand_day}"
            wb.save(NAME_FILE_RANDOM)
            print(f"File:{NAME_FILE_RANDOM} generated random - ok! ROWS:{COUNT_RECORDS}")


# перезапись и.или создание
# таблицы на основе каталога с изображениями
def generate_from_image(directory):
        files = os.listdir(path=directory)
        images = []
        for file in files:
            if str(file).endswith(".jpg"):
                   images.append(directory+"/"+file)

        names, values, dates = extr.launch_loop(images)

        wb = Workbook()
        table_xlsx = wb.active
        table_xlsx.title = NAME_FILE_IMAGES
        for i in range(len(names)-1):
            index = str(i+1)
            table_xlsx['A'+index] = names[1]
            table_xlsx['B'+index] = values[i]
            table_xlsx['C'+index] = dates[i]
        wb.save(NAME_FILE_IMAGES)
        print(f"File:{NAME_FILE_IMAGES} generated from images - ok! ROWS:{len(names)}")


# конвертирование строки в дату
def str_to_date( str_date):
    if type(str_date) is str:
        if str_date.find("/") != -1:
            str_date = str_date.split("/")
        elif str_date.find("-") != -1:
            str_date = str_date.split("-")
        str_date = f"{str_date[0]}-{str_date[1]}-{str_date[2]}"
        str_date = datetime.datetime.strptime(str_date, "%Y-%m-%d").date()
        return str_date
    return "error type date"

# сортировка xlsx таблицы по дате
def sorted_xlsx_dates():
        check_file = os.path.exists(NAME_FILE_SORTED)
        if check_file == False:
            print(f"Not Fonud DB in path: {NAME_FILE_SORTED}")
            return 

        wb_obj = load_workbook(NAME_FILE_SORTED)

        sheet_obj = wb_obj.active
        m_row = sheet_obj.max_row

        names, values, dates = [],[],[]
        for i in range(1, m_row + 1):
            n, v, d = sheet_obj.cell(row = i, column = 1),sheet_obj.cell(row = i, column = 2),sheet_obj.cell(row = i, column = 3)
            names.append(n.value)
            values.append(v.value)
            dates.append(d.value)
        
        swap_bool = True
        while swap_bool:
            swap_bool = False
            for i in range(len(dates) - 1):
                n1, n2 = dates[i], dates[i + 1]
                n1 = str_to_date(n1) 
                n2 = str_to_date(n2)
                if n1 > n2:
                    dates[i], dates[i + 1] = dates[i + 1], dates[i]
                    names[i], names[i + 1] = names[i + 1], names[i]
                    values[i], values[i + 1] = values[i + 1], values[i]
                    swap_bool = True
        wb = Workbook()
        table_xlsx = wb.active
        table_xlsx.title = NAME_FILE_SORTED
        for i in range(len(names)-1):
            index = str(i+1)
            table_xlsx['A'+index] = names[1]
            table_xlsx['B'+index] = values[i]
            table_xlsx['C'+index] = dates[i]
        wb.save(NAME_FILE_SORTED)
        print(f"File:{NAME_FILE_SORTED} sorted - ok! Rows:{len(names)}")


YOUR_PATH = "путь к папке с фотографиями"
if "путь к папке с фотографиями" == YOUR_PATH:
    print("Добавь свой путь к фотографиям")
    sys.exit(0)

generate_from_image(YOUR_PATH)
# generate_random_xlsx()
# sorted_xlsx_dates()