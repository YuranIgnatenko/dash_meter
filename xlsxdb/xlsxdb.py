from openpyxl import load_workbook
import os.path
from datetime import date
import datetime

class Controller():
    def __init__(self, namefile_db):
        self.namefile_db = namefile_db
        self.select_names = []
        self.select_values = []
        self.select_dates  = []

    # return date format (12-12-2000)
    # input '12/12/1212' "12-12-1212"
    def str_to_date(self, str_date):
        if type(str_date) is str:
            if str_date.find("/") != -1:
                str_date = str_date.split("/")
            elif str_date.find("-") != -1:
                str_date = str_date.split("-")
            str_date = f"{str_date[0]}-{str_date[1]}-{str_date[2]}"
            str_date = datetime.datetime.strptime(str_date, "%Y-%m-%d").date()
            return str_date
        return "error type date"


    def sorted_on_dates(self, names, values, dates):
        swap_bool = True
        while swap_bool:
            swap_bool = False
            for i in range(len(dates) - 1):
                n1, n2 = dates[i], dates[i + 1]
                n1 = self.str_to_date(n1) 
                n2 = self.str_to_date(n2)
                if n1 > n2:
                    dates[i], dates[i + 1] = dates[i + 1], dates[i]
                    names[i], names[i + 1] = names[i + 1], names[i]
                    values[i], values[i + 1] = values[i + 1], values[i]
                    swap_bool = True
        return names, values, dates
            

    def read_all_rows(self):
        check_file = os.path.exists(self.namefile_db)
        if check_file == False:
            print(f"Not Fonud DB in path: {self.namefile_db}")
            return [-1], [-1], [-1]

        wb_obj = load_workbook(self.namefile_db)

        sheet_obj = wb_obj.active
        m_row = sheet_obj.max_row

        names, values, dates = [],[],[]
        for i in range(1, m_row + 1):
            n, v, d = sheet_obj.cell(row = i, column = 1),sheet_obj.cell(row = i, column = 2),sheet_obj.cell(row = i, column = 3)
            names.append(n.value)
            values.append(v.value)
            dates.append(d.value)
        
        return self.sorted_on_dates(names, values, dates)

    
    # dates: '12-12-2000', '12/12/2000' 
    def read_slice_time(self, date_min, date_max):
        if type(date_min) is str:
            date_min = self.str_to_date(date_min)
            date_max = self.str_to_date(date_max)
        elif date_min[0] > 999:
            date_min = date(date_min[0], date_min[1], date_min[2])
            date_max = date(date_max[0], date_max[1], date_max[2])

        names, values, dates = self.read_all_rows()
        self.select_names = []
        self.select_values = []
        self.select_dates  = []

        if names[0] == -1:
            return False
    
        for i in range(len(names)):
            dt = dates[i]
            if dt.find("/") != -1:
                dt = dt.split("/")
            elif dt.find("-") != -1:
                dt = dt.split("-")

            dt = f"{dt[0]}-{dt[1]}-{dt[2]}"
            dt = datetime.datetime.strptime(dt, "%Y-%m-%d").date()
            

            if (date_min <= dt) and (dt <= date_max):
                self.select_names.append(names[i])
                self.select_values.append(values[i])
                self.select_dates.append(dates[i])
                
        return date_min, date_max